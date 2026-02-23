"""XLSX 导入路由"""
import hashlib
import json
import re
import uuid
from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, UploadFile, File, HTTPException
from openpyxl import load_workbook

from backend.config import SNAPSHOT_DIR
from backend.database import get_db
from backend.services.xlsx_parser import parse_xlsx_rows
from backend.services.diff_service import compute_diffs

router = APIRouter()


@router.post("/xlsx")
async def upload_xlsx(file: UploadFile = File(...), force: bool = False):
    """上传 xlsx 文件，解析并生成 diff

    force=true 时跳过重复检查（用于重新导入修正后的解析器）
    """
    if not file.filename or not file.filename.endswith('.xlsx'):
        raise HTTPException(400, "请上传 .xlsx 文件")

    content = await file.read()
    file_hash = hashlib.md5(content).hexdigest()

    db = await get_db()

    # 检查是否已导入过（force=true 时跳过）
    if not force:
        cursor = await db.execute(
            "SELECT id FROM xlsx_snapshots WHERE file_hash = ?", (file_hash,)
        )
        if await cursor.fetchone():
            raise HTTPException(409, "此文件已导入过（内容完全相同）。加 ?force=true 强制重新导入。")

    # 保存快照文件
    SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    snapshot_path = SNAPSHOT_DIR / f"{ts}_{file.filename}"
    snapshot_path.write_bytes(content)

    # 读取 xlsx (不用 read_only 避免兼容问题)
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    if ws is None:
        raise HTTPException(400, "xlsx 文件没有有效工作表")

    # 读取表头 + 数据行
    rows_data: list[dict] = []
    headers: list[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h or '').strip() for h in row]
            continue
        row_dict = {}
        for j, val in enumerate(row):
            if j < len(headers) and headers[j]:
                row_dict[headers[j]] = val
        if any(v for v in row_dict.values()):
            rows_data.append(row_dict)
    wb.close()

    # 默认年份（从文件名推断）
    default_year = datetime.now().year
    year_match = re.search(r'20\d{2}', file.filename or '')
    if year_match:
        default_year = int(year_match.group())

    # 解析
    parsed_rows = parse_xlsx_rows(rows_data, default_year)

    # 统计 fragment 数量
    total_fragments = sum(len(r.fragments) for r in parsed_rows)

    # 写入快照记录 (force 模式下用新 hash 避免 UNIQUE 冲突)
    actual_hash = file_hash if not force else f"{file_hash}_{ts}"
    cursor = await db.execute(
        "INSERT INTO xlsx_snapshots (file_name, file_hash, file_path, row_count) VALUES (?, ?, ?, ?)",
        (file.filename, actual_hash, str(snapshot_path), len(parsed_rows)),
    )
    snapshot_id = cursor.lastrowid

    # 计算 diff（在写入数据前）
    diffs = await compute_diffs(db, parsed_rows, snapshot_id)

    # 写入/更新 products 和 leads
    lead_inserted = 0
    for row in parsed_rows:
        p = row.product
        await db.execute("""
            INSERT INTO products (sku_code, product_id, product_name, price, shop_name,
                commission_rate, promo_link, launch_date, product_status, abnormal_note)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(sku_code) DO UPDATE SET
                product_id = excluded.product_id,
                product_name = excluded.product_name,
                price = excluded.price,
                shop_name = excluded.shop_name,
                commission_rate = excluded.commission_rate,
                promo_link = excluded.promo_link,
                launch_date = excluded.launch_date,
                product_status = excluded.product_status,
                abnormal_note = excluded.abnormal_note,
                updated_at = datetime('now','localtime')
        """, (
            p.sku_code, p.product_id, p.product_name, p.price, p.shop_name,
            p.commission_rate, p.promo_link, p.launch_date, p.product_status, p.abnormal_note,
        ))

        # 写入 listing 历史
        if p.promo_link or p.shop_name:
            # 查已有 active listing
            cursor = await db.execute(
                "SELECT id, shop_name, promo_link FROM listings WHERE sku_code = ? AND is_active = 1",
                (p.sku_code,),
            )
            existing_listing = await cursor.fetchone()
            if existing_listing:
                old_shop = existing_listing["shop_name"] or ""
                old_link = existing_listing["promo_link"] or ""
                if old_shop != p.shop_name or old_link != p.promo_link:
                    # 旧 listing 标记失活
                    await db.execute(
                        "UPDATE listings SET is_active = 0, end_date = datetime('now','localtime') WHERE id = ?",
                        (existing_listing["id"],),
                    )
                    # 新 listing
                    await db.execute(
                        "INSERT INTO listings (sku_code, shop_name, promo_link, product_id, snapshot_id) VALUES (?, ?, ?, ?, ?)",
                        (p.sku_code, p.shop_name, p.promo_link, p.product_id, snapshot_id),
                    )
            else:
                # 首次 listing
                await db.execute(
                    "INSERT INTO listings (sku_code, shop_name, promo_link, product_id, snapshot_id) VALUES (?, ?, ?, ?, ?)",
                    (p.sku_code, p.shop_name, p.promo_link, p.product_id, snapshot_id),
                )

        for frag in row.fragments:
            lead_id = str(uuid.uuid4())
            try:
                await db.execute("""
                    INSERT INTO leads (id, sku_code, snapshot_id, material_year,
                        raw_fragment, normalized_fragment, normalized_fragment_hash,
                        material_month, material_day, session_label, host_label,
                        time_points_json, parse_confidence)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    lead_id, p.sku_code, snapshot_id, frag.material_year,
                    frag.raw_fragment, frag.normalized_fragment, frag.normalized_fragment_hash,
                    frag.material_month, frag.material_day, frag.session_label, frag.host_label,
                    json.dumps(frag.time_points), frag.parse_confidence,
                ))
                lead_inserted += 1
            except Exception:
                pass  # UNIQUE 冲突跳过

    await db.commit()

    return {
        "snapshot_id": snapshot_id,
        "file_name": file.filename,
        "row_count": len(parsed_rows),
        "total_fragments": total_fragments,
        "lead_inserted": lead_inserted,
        "diff_count": len(diffs),
        "diffs": diffs,
    }


@router.get("/snapshots")
async def list_snapshots():
    """列出历史快照"""
    db = await get_db()
    cursor = await db.execute(
        "SELECT id, file_name, file_hash, row_count, imported_at FROM xlsx_snapshots ORDER BY id DESC"
    )
    rows = await cursor.fetchall()
    return [dict(r) for r in rows]


@router.get("/diffs/{snapshot_id}")
async def get_diffs(snapshot_id: int):
    """某次导入的变更明细"""
    db = await get_db()
    cursor = await db.execute(
        "SELECT id, snapshot_id, diff_type, sku_code, detail_json, created_at FROM import_diffs WHERE snapshot_id = ? ORDER BY id",
        (snapshot_id,),
    )
    rows = await cursor.fetchall()
    return [dict(r) for r in rows]
